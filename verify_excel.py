# coding: utf-8
import openpyxl

wb = openpyxl.load_workbook('Autotest/任务设计&检验逻辑.xlsx')
ws = wb.active

print(f"工作表: {ws.title}")
print(f"总行数: {ws.max_row}")
print(f"总列数: {ws.max_column}\n")

print("前5个任务:")
for row in range(2, 7):
    task_num = ws.cell(row=row, column=1).value
    instruction = ws.cell(row=row, column=2).value
    print(f"任务{task_num}: {instruction[:40]}...")

print(f"\n最后5个任务:")
for row in range(32, 37):
    task_num = ws.cell(row=row, column=1).value
    instruction = ws.cell(row=row, column=2).value
    print(f"任务{task_num}: {instruction[:40]}...")
