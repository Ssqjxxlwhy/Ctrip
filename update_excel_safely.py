import openpyxl
import os
import sys

# Excel文件路径
excel_path = r"Autotest\任务设计&检验逻辑.xlsx"

try:
    # 打开Excel文件
    print(f"正在打开Excel文件: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    print("Excel文件打开成功！")

    # 更新任务22的检验方法和检验脚本
    # 任务22在第23行（行号从1开始）
    print("\n更新任务22...")
    row_22 = 23
    ws.cell(row=row_22, column=3).value = """1. 计算真实答案：从北京酒店数据中按评分降序排序，取前5家
2. 智能体答案：接收智能体返回的酒店列表（可以是酒店名称列表或酒店ID列表）
3. 对比判断：将智能体答案与真实答案进行比较，前5家酒店名称必须完全一致且顺序相同
4. 辅助检查：也可从search_params.json检查是否记录了{type:"hotel_search", city:"北京", sortBy:"rating", limit:5, list_shown:true}"""
    ws.cell(row=row_22, column=4).value = "check_hotel_search_beijing_top5_rating(agent_answer)"

    # 更新任务23的检验方法和检验脚本
    # 任务23在第24行
    print("更新任务23...")
    row_23 = 24
    ws.cell(row=row_23, column=3).value = """1. 计算真实答案：从上海酒店数据中按价格升序排序，取第1家（价格最低）
2. 智能体答案：接收智能体返回的酒店（可以是酒店名称、酒店ID或包含酒店信息的字典）
3. 对比判断：将智能体答案与真实答案进行比较，必须是价格最低的那家酒店
4. 辅助检查：也可从search_params.json检查是否记录了{type:"hotel_search", city:"上海", sortBy:"price_asc", list_shown:true}"""
    ws.cell(row=row_23, column=4).value = "check_hotel_search_shanghai_cheapest(agent_answer)"

    # 更新任务24的检验方法和检验脚本
    # 任务24在第25行
    print("更新任务24...")
    row_24 = 25
    ws.cell(row=row_24, column=3).value = """1. 计算真实答案：
   - 根据日期计算基础价格：dayPrice = 380 - 10 = 370（10月21日是周一）
   - 生成北京到广州的所有航班数据（2个经济舱 + 2个公务/头等舱）
   - 按价格排序，取最便宜的3趟：370, 395, 1110
   - 计算平均价格：(370 + 395 + 1110) / 3 = 625元
2. 智能��答案：接收智能体返回的平均价格（可以是数字、字符串或包含平均价格的字典）
3. 对比判断：允许±1元的误差（考虑四舍五入）
4. 辅助检查：也可从search_params.json检查是否记录了{type:"flight_search", from:"北京", to:"广州", date:"2025-10-21", calculation:"average_price_top3", list_shown:true}"""
    ws.cell(row=row_24, column=4).value = "check_flight_search_price_avg(agent_answer)"

    # 更新任务25的检验方法和检验脚本
    # 任务25在第26行
    print("更新任务25...")
    row_25 = 26
    ws.cell(row=row_25, column=3).value = """1. 计算真实答案：
   - 从trains_list.json加载火车票数据
   - 筛选广州到杭州10月22日的所有车次
   - 筛选出发时间在14:00-17:00之间的车次
   - 真实答案：K535（14:00-19:50）
2. 智能体答案：接收智能体返回的车次列表（可以是车次号列表或包含车次信息的字典列表）
3. 对比判断：将智能体答案与真实答案进行比较，必须包含所有符合条件的车次
4. 辅助检查：也可从search_params.json检查是否记录了{type:"train_search", from:"广州", to:"杭州", date:"2025-10-22", timeRange:"14:00-17:00", list_shown:true}"""
    ws.cell(row=row_25, column=4).value = "check_train_search_time_range(agent_answer)"

    # 保存文件
    print("\n正在保存Excel文件...")
    wb.save(excel_path)
    print("Excel文件更新成功！")

    # 打印更新内容
    print("\n" + "=" * 80)
    print("更新内容摘要：")
    print("=" * 80)
    print(f"任务22 - 检验函数：check_hotel_search_beijing_top5_rating(agent_answer)")
    print(f"         检验方法：计算北京评分最高的前5家酒店，对比智能体答案")
    print()
    print(f"任务23 - 检验函数：check_hotel_search_shanghai_cheapest(agent_answer)")
    print(f"         检验方法：计算上海价格最低的酒店，对比智能体答案")
    print()
    print(f"任务24 - 检验函数：check_flight_search_price_avg(agent_answer)")
    print(f"         检验方法：计算北京到广州最便宜3趟航班的平均价格(625元)，对比智能体答案")
    print()
    print(f"任务25 - 检验函数：check_train_search_time_range(agent_answer)")
    print(f"         检验方法：计算广州到杭州14:00-17:00的车次(K535)，对比智能体答案")
    print("=" * 80)

except PermissionError:
    print("错误：Excel文件可能正在被其他程序打开。")
    print("请关闭Excel文件后重新运行此脚本。")
    sys.exit(1)
except Exception as e:
    print(f"错误：{e}")
    sys.exit(1)
