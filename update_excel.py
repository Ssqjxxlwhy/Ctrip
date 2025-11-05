# coding: utf-8
import openpyxl
import json

# 任务数据存在JSON中避免编码问题
tasks_data = '''[
    ["1", "点击 \\"酒店\\" 图标，进入酒店预订页面。", "检查click_history.json中的点击记录是否为 \\"点击'酒店'图标并进入酒店预订页面\\" ", "check_click_hotel()"],
    ["2", "点击 \\"机票\\" 图标，进入机票预订界面。", "检查click_history.json中的点击记录是否为 \\"点击'机票'图标并进入机票预订界面\\" ", "check_click_flight()"],
    ["3", "点击 \\"火车票\\" 图标，进入火车票预订页面。", "检查click_history.json中的点击记录是否为 \\"点击'火车票'图标并进入火车票预订页面\\" ", "check_click_train()"],
    ["4", "点击\\"消息\\"按钮，进入消息页面", "检查click_history.json中的点击记录是否为 \\"点击'消息'按钮并进入消息页面\\" ", "check_click_message()"],
    ["5", "点击\\"行程\\"按钮，进入行程页面", "检查click_history.json中的点击记录是否为 \\"点击'行程'按钮并进入行程页面\\" ", "check_click_itinerary()"],
    ["6", "点击\\"我的\\"按钮，进入我的页面", "检查click_history.json中的点击记录是否为 \\"点击'我的'按钮并进入我的页面\\" ", "check_click_profile()"],
    ["7", "进入 \\"酒店\\" 页面，选择城市成都，得到酒店列表", "检查search_params.json中的搜索记录是否包含{type:\\"hotel_search\\", city:\\"成都\\"}", "check_hotel_search_chengdu()"],
    ["8", "进入 \\"酒店\\" 页面，入住时间选择10月20日，退房时间选择10月21日，得到酒店列表", "检查search_params.json中的搜索记录是否包含{type:\\"hotel_search\\", checkIn:\\"2025-10-20\\", checkOut:\\"2025-10-21\\"}", "check_hotel_search_dates()"],
    ["9", "进入 \\"酒店\\" 页面，选择间数1、成人数1、儿童数1，得到酒店列表", "检查search_params.json中的搜索记录是否包含{type:\\"hotel_search\\", rooms:1, adults:1, children:1}", "check_hotel_search_guests()"],
    ["10", "进入 \\"机票\\" 页面，选出发地 \\"广州\\"、目的地 \\"深圳\\"，得到航班列表", "检查search_params.json中的搜索记录是否包含{type:\\"flight_search\\", from:\\"广州\\", to:\\"深圳\\"}", "check_flight_search_gz_sz()"],
    ["11", "进入 \\"机票\\" 页面，选择日期10月22日，得到航班列表", "检查search_params.json中的搜索记录是否包含{type:\\"flight_search\\", date:\\"2025-10-22\\"}", "check_flight_search_date()"],
    ["12", "进入\\"机票\\"页面，选择经济舱或者公务/头等舱，得到航班列表", "检查search_params.json中的搜索记录cabin字段是否为\\"经济舱\\"或\\"公务舱\\"或\\"头等舱\\"", "check_flight_search_cabin()"],
    ["13", "进入 \\"火车票\\" 页面，选出发地 \\"北京\\"、目的地 \\"上海\\"，得到车次列表", "检查search_params.json中的搜索记录是否包含{type:\\"train_search\\", from:\\"北京\\", to:\\"上海\\"}", "check_train_search_bj_sh()"],
    ["14", "进入 \\"火车票\\" 页面，选择日期10月24日，得到车次列表", "检查search_params.json中的搜索记录是否包含{type:\\"train_search\\", date:\\"2025-10-24\\"}", "check_train_search_date()"],
    ["15", "进入 \\"火车票\\" 页面，选择学生票，得到车次列表", "检查search_params.json中的搜索记录是否包含{type:\\"train_search\\", ticketType:\\"学生票\\"}", "check_train_search_student()"],
    ["16", "进入 \\"酒店\\" 页面，城市选北京，入住时间选择10月22日，退房时间选择10月23日，得到酒店列表", "检查search_params.json中的搜索记录是否包含{type:\\"hotel_search\\", city:\\"北京\\", checkIn:\\"2025-10-22\\", checkOut:\\"2025-10-23\\"}", "check_hotel_search_beijing()"],
    ["17", "进入 \\"酒店\\" 页面，城市选上海，入住时间今天，退房时间明天，间数选择2，成人数选择2，儿童数选择0，得到酒店列表", "检查search_params.json中的搜索记录是否包含{type:\\"hotel_search\\", city:\\"上海\\", rooms:2, adults:2, children:0}（日期动态）", "check_hotel_search_shanghai()"],
    ["18", "进入 \\"机票\\" 页面，选出发地 \\"北京\\"、目的地 \\"深圳\\"，选择日期10月25号，得到航班列表", "检查search_params.json中的搜索记录是否包含{type:\\"flight_search\\", from:\\"北京\\", to:\\"深圳\\", date:\\"2025-10-25\\"}", "check_flight_search_bj_sz()"],
    ["19", "进入 \\"机票\\" 页面，选出发地 \\"成都\\"、目的地 \\"上海\\"，选择日期10月20号，选头等舱，得到航班列表", "检查search_params.json中的搜索记录是否包含{type:\\"flight_search\\", from:\\"成都\\", to:\\"上海\\", date:\\"2025-10-20\\", cabin:\\"头等舱\\"}", "check_flight_search_cd_sh_first()"],
    ["20", "进入 \\"火车票\\" 页面，选出发地 \\"北京\\"、目的地 \\"上海\\"，选择日期10月22日，得到车次列表", "检查search_params.json中的搜索记录是否包含{type:\\"train_search\\", from:\\"北京\\", to:\\"上海\\", date:\\"2025-10-22\\"}", "check_train_search_bj_sh_date()"],
    ["21", "进入 \\"火车票\\" 页面，选出发地 \\"长沙\\"、目的地 \\"天津\\"，选择日期1月18日，选择学生票，得到车次列表", "检查search_params.json中的搜索记录是否包含{type:\\"train_search\\", from:\\"长沙\\", to:\\"天津\\", date:\\"2026-01-18\\", ticketType:\\"学生票\\"}", "check_train_search_cs_tj_student()"],
    ["22", "进入 \\"酒店\\" 页面，为我筛选北京评分最高的前 5 家酒店", "检查search_params.json中的搜索记录是否包含{type:\\"hotel_search\\", city:\\"北京\\", sortBy:\\"rating\\", limit:5}", "check_hotel_search_beijing_top5()"],
    ["23", "进入 \\"酒店\\" 页面，为我检索上海价格最低的酒店", "检查search_params.json中的搜索记录是否包含{type:\\"hotel_search\\", city:\\"上海\\", sortBy:\\"price_asc\\"}", "check_hotel_search_shanghai_cheapest()"],
    ["24", "进入 \\"机票\\" 页面，查10月21日从北京飞广州的机票，统计下最便宜的 3 趟航班的平均价格。", "检查search_params.json中的搜索记录是否包含{type:\\"flight_search\\", from:\\"北京\\", to:\\"广州\\", date:\\"2025-10-21\\", calculation:\\"average_price_top3\\"}", "check_flight_search_price_avg()"],
    ["25", "进入 \\"火车票\\" 页面，查10月22日广州到杭州下午2点到5点的车次，得到车次列表。", "检查search_params.json中的搜索记录是否包含{type:\\"train_search\\", from:\\"广州\\", to:\\"杭州\\", date:\\"2025-10-22\\", timeRange:\\"14:00-17:00\\"}", "check_train_search_time_range()"],
    ["26", "选上海的酒店，入住日期选10月22日，退房10月25日，客房数和入住人数默认，得到酒店列表后，点击第一个酒店，然后选择第一个房型预订", "检查booking_history.json中的预订记录是否包含{type:\\"hotel_booking\\", city:\\"上海\\", checkIn:\\"2025-10-22\\", checkOut:\\"2025-10-25\\", hotelIndex:0, roomIndex:0}", "check_booking_hotel_shanghai()"],
    ["27", "进入 \\"机票\\" 页面，选出发地 \\"武汉\\"、目的地 \\"深圳\\"，选择日期11月10号，舱型默认，得到航班列表后，预订第一架航班", "检查booking_history.json中的预订记录是否包含{type:\\"flight_booking\\", from:\\"武汉\\", to:\\"深圳\\", date:\\"2025-11-10\\", flightIndex:0}", "check_booking_flight_wh_sz()"],
    ["28", "进入 \\"火车票\\" 页面，选出发地 \\"北京\\"、目的地 \\"上海\\"，选择日期10月23日，得到车次列表后，预订第一班车次", "检查booking_history.json中的预订记录是否包含{type:\\"train_booking\\", from:\\"北京\\", to:\\"上海\\", date:\\"2025-10-23\\", trainIndex:0}", "check_booking_train_bj_sh()"],
    ["29", "选上海的酒店，入住日期选10月21日，退房10月25日，客房数和入住人数默认，得到酒店列表后，选择价格最低的酒店，然后选择最便宜的房型预订", "检查booking_history.json中的预订记录是否包含{type:\\"hotel_booking\\", city:\\"上海\\", checkIn:\\"2025-10-21\\", checkOut:\\"2025-10-25\\", selection:\\"cheapest\\"}", "check_booking_hotel_cheapest()"],
    ["30", "进入 \\"火车票\\" 页面，选出发地 \\"北京\\"、目的地 \\"上海\\"，选择日期10月20日，得到车次列表后，预订下午3点到下午4点的任意一班车次", "检查booking_history.json中的预订记录是否包含{type:\\"train_booking\\", from:\\"北京\\", to:\\"上海\\", date:\\"2025-10-20\\", timeRange:\\"15:00-16:00\\"}", "check_booking_train_time()"],
    ["31", "先订北京飞上海的机票（10月20日，经济舱），再订上海的酒店（入住10月20日，退房10月21日），最后订上海回北京的火车票（10月21日）", "检查booking_history.json中是否包含type:\\"multi_booking\\"且有3个步骤：机票(北京->上海)、酒店(上海)、火车票(上海->北京)，并验证顺序", "check_booking_multi_step()"],
    ["32", "订 10 月 20 日从杭州到北京的最快火车票（5 小时内达），住北京中关村亚朵酒店两晚（10.20 - 10.22），再订 10.22 北京回杭州的最晚高铁，计算所有费用后判断 2000 元够不够。", "检查booking_history.json中是否包含type:\\"complex_booking\\"且有：火车票(杭州->北京,<=5h)、酒店(中关村亚朵,2晚)、火车票(北京->杭州,最晚)、totalCost、budgetCheck:2000", "check_booking_complex_budget()"],
    ["33", "订 5 张 10 月 20 日深圳到北京的火车票（要求 5 小时内到达）", "检查booking_history.json中是否包含{type:\\"batch_booking\\", bookingType:\\"train\\", from:\\"深圳\\", to:\\"北京\\", date:\\"2025-10-20\\", quantity:5, constraint:\\"duration<=5h\\"}", "check_booking_batch_train()"],
    ["34", "订 5 张 10 月 20 日广州到北京的最早火车票，再订北京 3 家不同的酒店", "检查booking_history.json中是否包含type:\\"complex_batch_booking\\"且有：trainBooking(5张,最早) 和 hotelBooking(3家,不同)", "check_booking_complex_batch()"],
    ["35", "订 10 月 15 日北京飞上海的 6 张经济舱机票", "检查booking_history.json中是否包含{type:\\"batch_booking\\", bookingType:\\"flight\\", from:\\"北京\\", to:\\"上海\\", date:\\"2025-10-15\\", cabin:\\"经济舱\\", quantity:6}", "check_booking_batch_flight()"]
]'''

# 解析JSON数据
tasks = json.loads(tasks_data)

# 读取现有Excel文件
wb = openpyxl.load_workbook('Autotest/任务设计&检验逻辑.xlsx')
ws = wb.active

# 填充数据，从第2行开始
for idx, task in enumerate(tasks, start=2):
    ws.cell(row=idx, column=1, value=task[0])
    ws.cell(row=idx, column=2, value=task[1])
    ws.cell(row=idx, column=3, value=task[2])
    ws.cell(row=idx, column=4, value=task[3])

# 保存
wb.save('Autotest/任务设计&检验逻辑.xlsx')
print('Success! Excel file updated.')
print('File: Autotest/任务设计&检验逻辑.xlsx')
print('Updated 35 tasks')
