import openpyxl
import sys
import json

# 读取Autotest目录下的Excel文件
excel_path = r"Autotest\任务设计&检验逻辑.xlsx"

try:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    tasks = []

    # 读取任务22-25的内容
    for row_num in range(23, 27):  # 第23-26行对应任务22-25
        task_num = ws.cell(row=row_num, column=1).value
        instruction = ws.cell(row=row_num, column=2).value
        verification_method = ws.cell(row=row_num, column=3).value
        verification_function = ws.cell(row=row_num, column=4).value

        tasks.append({
            "task_num": task_num,
            "instruction": instruction,
            "verification_method": verification_method,
            "verification_function": verification_function
        })

    # 输出为JSON格式保存到文件
    with open("autotest_tasks_22_25.json", "w", encoding="utf-8") as f:
        json.dump(tasks, f, ensure_ascii=False, indent=2)

    print("成功读取任务22-25的内容，已保存到 autotest_tasks_22_25.json")

except Exception as e:
    print(f"错误: {e}")
    sys.exit(1)
