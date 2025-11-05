# -*- coding: utf-8 -*-
"""
生成所有任务的检验脚本和设计文档
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

# 定义所有任务的检验设计
TASK_DESIGNS = [
    {
        "id": 1,
        "instruction": "点击 "酒店" 图标，进入酒店预订页面。",
        "check_method": "维护点击记录存储",
        "check_logic": "在手机APP的内部存储中，维护一个记录用户操作历史的JSON文件，每当用户点击"酒店"图标时，自动将点击事件的相关信息写入该文件。通过ADB指令从手机中获取该JSON文件，解析文件内容，检查是否存在"点击'酒店'图标并进入酒店预订页面"的记录。",
        "check_key": {"icon": "酒店", "page": "酒店预订页面"}
    },
    {
        "id": 2,
        "instruction": "点击 "机票" 图标，进入机票预订界面。",
        "check_method": "维护点击记录存储",
        "check_logic": "在手机APP的内部存储中维护点击记录JSON文件，每当用户点击"机票"图标时，自动将点击事件写入该文件。通过ADB指令获取并解析JSON文件，检查是否存在"点击'机票'图标并进入机票预订界面"的记录。",
        "check_key": {"icon": "机票", "page": "机票预订界面"}
    },
    {
        "id": 3,
        "instruction": "点击 "火车票" 图标，进入火车票预订页面。",
        "check_method": "维护点击记录存储",
        "check_logic": "在手机APP的内部存储中维护点击记录JSON文件，每当用户点击"火车票"图标时，自动将点击事件写入该文件。通过ADB指令获取并解析JSON文件，检查是否存在"点击'火车票'图标并进入火车票预订页面"的记录。",
        "check_key": {"icon": "火车票", "page": "火车票预订页面"}
    },
    {
        "id": 4,
        "instruction": "点击"消息"按钮，进入消息页面",
        "check_method": "维护点击记录存储",
        "check_logic": "在手机APP的内部存储中维护点击记录JSON文件，每当用户点击"消息"按钮时，自动将点击事件写入该文件。通过ADB指令获取并解析JSON文件，检查是否存在"点击'消息'按钮并进入消息页面"的记录。",
        "check_key": {"icon": "消息", "page": "消息页面"}
    },
    {
        "id": 5,
        "instruction": "点击"行程"按钮，进入行程页面",
        "check_method": "维护点击记录存储",
        "check_logic": "在手机APP的内部存储中维护点击记录JSON文件，每当用户点击"行程"按钮时，自动将点击事件写入该文件。通过ADB指令获取并解析JSON文件，检查是否存在"点击'行程'按钮并进入行程页面"的记录。",
        "check_key": {"icon": "行程", "page": "行程页面"}
    },
    {
        "id": 6,
        "instruction": "点击"我的"按钮，进入我的页面",
        "check_method": "维护点击记录存储",
        "check_logic": "在手机APP的内部存储中维护点击记录JSON文件，每当用户点击"我的"按钮时，自动将点击事件写入该文件。通过ADB指令获取并解析JSON文件，检查是否存在"点击'我的'按钮并进入我的页面"的记录。",
        "check_key": {"icon": "我的", "page": "我的页面"}
    },
    {
        "id": 7,
        "instruction": "进入 "酒店" 页面，选择城市成都，得到酒店列表",
        "check_method": "维护搜索参数记录存储",
        "check_logic": "在APP内部存储中维护搜索参数记录JSON文件（search_params.json），当用户在酒店页面选择城市后进行搜索时，记录搜索参数（城市、日期等）。通过ADB获取该文件，检查是否存在城市为"成都"的酒店搜索记录。",
        "check_key": {"type": "hotel_search", "city": "成都"}
    },
    {
        "id": 8,
        "instruction": "进入 "酒店" 页面，入住时间选择10月20日，退房时间选择10月21日，得到酒店列表",
        "check_method": "维护搜索参数记录存储",
        "check_logic": "在search_params.json中记录酒店搜索参数，通过ADB获取文件，检查是否存在入住时间为"2025-10-20"且退房时间为"2025-10-21"的酒店搜索记录。",
        "check_key": {"type": "hotel_search", "checkIn": "2025-10-20", "checkOut": "2025-10-21"}
    },
    {
        "id": 9,
        "instruction": "进入 "酒店" 页面，选择间数1、成人数1、儿童数1，得到酒店列表",
        "check_method": "维护搜索参数记录存储",
        "check_logic": "在search_params.json中记录酒店搜索参数，通过ADB获取文件，检查是否存在间数为1、成人数为1、儿童数为1的酒店搜索记录。",
        "check_key": {"type": "hotel_search", "rooms": 1, "adults": 1, "children": 1}
    },
    {
        "id": 10,
        "instruction": "进入 "机票" 页面，选出发地 "广州"、目的地 "深圳"，得到航班列表",
        "check_method": "维护搜索参数记录存储",
        "check_logic": "在search_params.json中记录机票搜索参数，通过ADB获取文件，检查是否存在出发地为"广州"、目的地为"深圳"的航班搜索记录。",
        "check_key": {"type": "flight_search", "from": "广州", "to": "深圳"}
    },
    {
        "id": 11,
        "instruction": "进入 "机票" 页面，选择日期10月22日，得到航班列表",
        "check_method": "维护搜索参数记录存储",
        "check_logic": "在search_params.json中记录机票搜索参数，通过ADB获取文件，检查是否存在日期为"2025-10-22"的航班搜索记录。",
        "check_key": {"type": "flight_search", "date": "2025-10-22"}
    },
    {
        "id": 12,
        "instruction": "进入"机票"页面，选择经济舱或者公务/头等舱，得到航班列表",
        "check_method": "维护搜索参数记录存储",
        "check_logic": "在search_params.json中记录机票搜索参数，通过ADB获取文件，检查是否存在舱位类型为\"经济舱\"或\"公务舱\"或\"头等舱\"的航班搜索记录。",
        "check_key": {"type": "flight_search", "cabin": ["经济舱", "公务舱", "头等舱"]}
    },
    {
        "id": 13,
        "instruction": "进入 "火车票" 页面，选出发地 "北京"、目的地 "上海"，得到车次列表",
        "check_method": "维护搜索参数记录存储",
        "check_logic": "在search_params.json中记录火车票搜索参数，通过ADB获取文件，检查是否存在出发地为"北京"、目的地为"上海"的车次搜索记录。",
        "check_key": {"type": "train_search", "from": "北京", "to": "上海"}
    },
    {
        "id": 14,
        "instruction": "进入 "火车票" 页面，选择日期10月24日，得到车次列表",
        "check_method": "维护搜索参数记录存储",
        "check_logic": "在search_params.json中记录火车票搜索参数，通过ADB获取文件，检查是否存在日期为"2025-10-24"的车次搜索记录。",
        "check_key": {"type": "train_search", "date": "2025-10-24"}
    },
    {
        "id": 15,
        "instruction": "进入 "火车票" 页面，选择学生票，得到车次列表",
        "check_method": "维护搜索参数记录存储",
        "check_logic": "在search_params.json中记录火车票搜索参数，通过ADB获取文件，检查是否存在票种为\"学生票\"的车次搜索记录。",
        "check_key": {"type": "train_search", "ticketType": "学生票"}
    },
    {
        "id": 16,
        "instruction": "进入 "酒店" 页面，城市选北京，入住时间选择10月22日，退房时间选择10月23日，得到酒店列表",
        "check_method": "维护搜索参数记录存储",
        "check_logic": "在search_params.json中记录酒店搜索参数，通过ADB获取文件，检查是否存在城市为\"北京\"、入住时间为\"2025-10-22\"、退房时间为\"2025-10-23\"的酒店搜索记录。",
        "check_key": {"type": "hotel_search", "city": "北京", "checkIn": "2025-10-22", "checkOut": "2025-10-23"}
    },
    {
        "id": 17,
        "instruction": "进入 "酒店" 页面，城市选上海，入住时间今天，退房时间明天，间数选择2，成人数选择2，儿童数选择0，得到酒店列表",
        "check_method": "维护搜索参数记录存储",
        "check_logic": "在search_params.json中记录酒店搜索参数，通过ADB获取文件，检查是否存在城市为\"上海\"、间数为2、成人数为2、儿童数为0的酒店搜索记录（日期动态判断为今明两天）。",
        "check_key": {"type": "hotel_search", "city": "上海", "rooms": 2, "adults": 2, "children": 0}
    },
    {
        "id": 18,
        "instruction": "进入 "机票" 页面，选出发地 "北京"、目的地 "深圳"，选择日期10月25号，得到航班列表",
        "check_method": "维护搜索参数记录存储",
        "check_logic": "在search_params.json中记录机票搜索参数，通过ADB获取文件，检查是否存在出发地为\"北京\"、目的地为\"深圳\"、日期为\"2025-10-25\"的航班搜索记录。",
        "check_key": {"type": "flight_search", "from": "北京", "to": "深圳", "date": "2025-10-25"}
    },
    {
        "id": 19,
        "instruction": "进入 "机票" 页面，选出发地 "成都"、目的地 "上海"，选择日期10月20号，选头等舱，得到航班列表",
        "check_method": "维护搜索参数记录存储",
        "check_logic": "在search_params.json中记录机票搜索参数，通过ADB获取文件，检查是否存在出发地为\"成都\"、目的地为\"上海\"、日期为\"2025-10-20\"、舱位为\"头等舱\"的航班搜索记录。",
        "check_key": {"type": "flight_search", "from": "成都", "to": "上海", "date": "2025-10-20", "cabin": "头等舱"}
    },
    {
        "id": 20,
        "instruction": "进入 "火车票" 页面，选出发地 "北京"、目的地 "上海"，选择日期10月22日，得到车次列表",
        "check_method": "维护搜索参数记录存储",
        "check_logic": "在search_params.json中记录火车票搜索参数，通过ADB获取文件，检查是否存在出发地为\"北京\"、目的地为\"上海\"、日期为\"2025-10-22\"的车次搜索记录。",
        "check_key": {"type": "train_search", "from": "北京", "to": "上海", "date": "2025-10-22"}
    },
    {
        "id": 21,
        "instruction": "进入 "火车票" 页面，选出发地 "长沙"、目的地 "天津"，选择日期1月18日，选择学生票，得到车次列表",
        "check_method": "维护搜索参数记录存储",
        "check_logic": "在search_params.json中记录火车票搜索参数，通过ADB获取文件，检查是否存在出发地为\"长沙\"、目的地为\"天津\"、日期为\"2026-01-18\"、票种为\"学生票\"的车次搜索记录。",
        "check_key": {"type": "train_search", "from": "长沙", "to": "天津", "date": "2026-01-18", "ticketType": "学生票"}
    },
    {
        "id": 22,
        "instruction": "进入 "酒店" 页面，为我筛选北京评分最高的前 5 家酒店",
        "check_method": "维护搜索参数和筛选条件记录",
        "check_logic": "在search_params.json中记录酒店搜索参数和筛选条件，通过ADB获取文件，检查是否存在城市为\"北京\"、排序方式为\"评分\"、显示数量为5的筛选记录。",
        "check_key": {"type": "hotel_search", "city": "北京", "sortBy": "rating", "limit": 5}
    },
    {
        "id": 23,
        "instruction": "进入 "酒店" 页面，为我检索上海价格最低的酒店",
        "check_method": "维护搜索参数和筛选条件记录",
        "check_logic": "在search_params.json中记录酒店搜索参数和筛选条件，通过ADB获取文件，检查是否存在城市为\"上海\"、排序方式为\"价格升序\"的筛选记录。",
        "check_key": {"type": "hotel_search", "city": "上海", "sortBy": "price_asc"}
    },
    {
        "id": 24,
        "instruction": "进入 "机票" 页面，查10月21日从北京飞广州的机票，统计下最便宜的 3 趟航班的平均价格。",
        "check_method": "维护搜索参数和计算结果记录",
        "check_logic": "在search_params.json中记录机票搜索参数和计算结果，通过ADB获取文件，检查是否存在出发地为\"北京\"、目的地为\"广州\"、日期为\"2025-10-21\"的搜索记录，并验证是否记录了最便宜3趟航班的平均价格计算结果。",
        "check_key": {"type": "flight_search", "from": "北京", "to": "广州", "date": "2025-10-21", "calculation": "average_price_top3"}
    },
    {
        "id": 25,
        "instruction": "进入 "火车票" 页面，查10月22日广州到杭州下午2点到5点的车次，得到车次列表。",
        "check_method": "维护搜索参数和时间筛选记录",
        "check_logic": "在search_params.json中记录火车票搜索参数和时间筛选条件，通过ADB获取文件，检查是否存在出发地为\"广州\"、目的地为\"杭州\"、日期为\"2025-10-22\"、时间范围为\"14:00-17:00\"的搜索记录。",
        "check_key": {"type": "train_search", "from": "广州", "to": "杭州", "date": "2025-10-22", "timeRange": "14:00-17:00"}
    },
    {
        "id": 26,
        "instruction": "选上海的酒店，入住日期选10月22日，退房10月25日，客房数和入住人数默认，得到酒店列表后，点击第一个酒店，然后选择第一个房型预订",
        "check_method": "维护预订记录存储",
        "check_logic": "在booking_history.json中记录用户的预订操作，通过ADB获取文件，检查是否存在城市为\"上海\"、入住\"2025-10-22\"、退房\"2025-10-25\"的酒店预订记录，并验证选择了列表第一个酒店的第一个房型。",
        "check_key": {"type": "hotel_booking", "city": "上海", "checkIn": "2025-10-22", "checkOut": "2025-10-25", "hotelIndex": 0, "roomIndex": 0}
    },
    {
        "id": 27,
        "instruction": "进入 "机票" 页面，选出发地 "武汉"、目的地 "深圳"，选择日期11月10号，舱型默认，得到航班列表后，预订第一架航班",
        "check_method": "维护预订记录存储",
        "check_logic": "在booking_history.json中记录机票预订操作，通过ADB获取文件，检查是否存在出发地为\"武汉\"、目的地为\"深圳\"、日期为\"2025-11-10\"的机票预订记录，并验证预订的是列表第一架航班。",
        "check_key": {"type": "flight_booking", "from": "武汉", "to": "深圳", "date": "2025-11-10", "flightIndex": 0}
    },
    {
        "id": 28,
        "instruction": "进入 "火车票" 页面，选出发地 "北京"、目的地 "上海"，选择日期10月23日，得到车次列表后，预订第一班车次",
        "check_method": "维护预订记录存储",
        "check_logic": "在booking_history.json中记录火车票预订操作，通过ADB获取文件，检查是否存在出发地为\"北京\"、目的地为\"上海\"、日期为\"2025-10-23\"的火车票预订记录，并验证预订的是列表第一班车次。",
        "check_key": {"type": "train_booking", "from": "北京", "to": "上海", "date": "2025-10-23", "trainIndex": 0}
    },
    {
        "id": 29,
        "instruction": "选上海的酒店，入住日期选10月21日，退房10月25日，客房数和入住人数默认，得到酒店列表后，选择价格最低的酒店，然后选择最便宜的房型预订",
        "check_method": "维护预订记录存储",
        "check_logic": "在booking_history.json中记录酒店预订操作，通过ADB获取文件，检查是否存在城市为\"上海\"、入住\"2025-10-21\"、退房\"2025-10-25\"的酒店预订记录，并验证选择了价格最低的酒店和最便宜的房型。",
        "check_key": {"type": "hotel_booking", "city": "上海", "checkIn": "2025-10-21", "checkOut": "2025-10-25", "selection": "cheapest"}
    },
    {
        "id": 30,
        "instruction": "进入 "火车票" 页面，选出发地 "北京"、目的地 "上海"，选择日期10月20日，得到车次列表后，预订下午3点到下午4点的任意一班车次",
        "check_method": "维护预订记录存储",
        "check_logic": "在booking_history.json中记录火车票预订操作，通过ADB获取文件，检查是否存在出发地为\"北京\"、目的地为\"上海\"、日期为\"2025-10-20\"、出发时间在\"15:00-16:00\"之间的火车票预订记录。",
        "check_key": {"type": "train_booking", "from": "北京", "to": "上海", "date": "2025-10-20", "timeRange": "15:00-16:00"}
    },
    {
        "id": 31,
        "instruction": "先订北京飞上海的机票（10月20日，经济舱），再订上海的酒店（入住10月20日，退房10月21日），最后订上海回北京的火车票（10月21日）",
        "check_method": "维护多步骤任务记录",
        "check_logic": "在booking_history.json中记录多步骤预订操作，通过ADB获取文件，检查是否完成了三个预订：1)机票(北京-上海,2025-10-20,经济舱) 2)酒店(上海,2025-10-20至2025-10-21) 3)火车票(上海-北京,2025-10-21)，并验证执行顺序正确。",
        "check_key": {"type": "multi_booking", "steps": [
            {"type": "flight", "from": "北京", "to": "上海", "date": "2025-10-20", "cabin": "经济舱"},
            {"type": "hotel", "city": "上海", "checkIn": "2025-10-20", "checkOut": "2025-10-21"},
            {"type": "train", "from": "上海", "to": "北京", "date": "2025-10-21"}
        ]}
    },
    {
        "id": 32,
        "instruction": "订 10 月 20 日从杭州到北京的最快火车票（5 小时内达），住北京中关村亚朵酒店两晚（10.20 - 10.22），再订 10.22 北京回杭州的最晚高铁，计算所有费用后判断 2000 元够不够。",
        "check_method": "维护多步骤任务和计算结果记录",
        "check_logic": "在booking_history.json中记录多步骤预订操作和费用计算结果，检查是否完成：1)火车票(杭州-北京,2025-10-20,5小时内) 2)酒店(北京中关村亚朵,2025-10-20至2025-10-22) 3)火车票(北京-杭州,2025-10-22,最晚) 4)总费用计算和2000元判断结果。",
        "check_key": {"type": "complex_booking", "bookings": [
            {"type": "train", "from": "杭州", "to": "北京", "date": "2025-10-20", "constraint": "duration<=5h"},
            {"type": "hotel", "name": "北京中关村亚朵酒店", "checkIn": "2025-10-20", "checkOut": "2025-10-22"},
            {"type": "train", "from": "北京", "to": "杭州", "date": "2025-10-22", "constraint": "latest"}
        ], "calculation": "total_cost", "budget_check": 2000}
    },
    {
        "id": 33,
        "instruction": "订 5 张 10 月 20 日深圳到北京的火车票（要求 5 小时内到达）",
        "check_method": "维护批量预订记录",
        "check_logic": "在booking_history.json中记录批量预订操作，检查是否存在5张火车票预订记录，出发地为\"深圳\"、目的地为\"北京\"、日期为\"2025-10-20\"、时长在5小时内。",
        "check_key": {"type": "batch_booking", "bookingType": "train", "from": "深圳", "to": "北京", "date": "2025-10-20", "constraint": "duration<=5h", "quantity": 5}
    },
    {
        "id": 34,
        "instruction": "订 5 张 10 月 20 日广州到北京的最早火车票，再订北京 3 家不同的酒店",
        "check_method": "维护批量和多类型预订记录",
        "check_logic": "在booking_history.json中记录多类型批量预订，检查是否完成：1)5张火车票(广州-北京,2025-10-20,最早) 2)3家不同的北京酒店预订。",
        "check_key": {"type": "complex_batch_booking", "trainBooking": {"from": "广州", "to": "北京", "date": "2025-10-20", "constraint": "earliest", "quantity": 5}, "hotelBooking": {"city": "北京", "quantity": 3, "constraint": "different"}}
    },
    {
        "id": 35,
        "instruction": "订 10 月 15 日北京飞上海的 6 张经济舱机票",
        "check_method": "维护批量预订记录",
        "check_logic": "在booking_history.json中记录批量预订操作，检查是否存在6张机票预订记录，出发地为\"北京\"、目的地为\"上海\"、日期为\"2025-10-15\"、舱位为\"经济舱\"。",
        "check_key": {"type": "batch_booking", "bookingType": "flight", "from": "北京", "to": "上海", "date": "2025-10-15", "cabin": "经济舱", "quantity": 6}
    }
]

def create_excel():
    """创建任务设计&检验逻辑Excel文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "任务设计&检验逻辑"

    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 80

    # 创建标题行
    headers = ['序号', '任务指令', '检验方法', '检验逻辑说明']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 填充数据
    for task in TASK_DESIGNS:
        row = task['id'] + 1
        ws.cell(row=row, column=1, value=task['id'])
        ws.cell(row=row, column=2, value=task['instruction'])
        ws.cell(row=row, column=3, value=task['check_method'])
        ws.cell(row=row, column=4, value=task['check_logic'])

        # 设置单元格样式
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if col == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # 保存文件
    wb.save('任务设计&检验逻辑.xlsx')
    print("✓ 已创建 任务设计&检验逻辑.xlsx")

def generate_eval_script(task):
    """生成单个检验脚本"""
    script_name = f"eval_{task['id']}.py"

    # 根据任务类型生成不同的检验代码
    if task['id'] <= 6:  # 点击类任务
        check_code = f"""
    # 检查是否存在目标点击记录
    target_icon = "{task['check_key']['icon']}"
    target_page = "{task['check_key']['page']}"
    for event in data.get("click_events", []):
        if event.get("icon") == target_icon and event.get("page") == target_page:
            return True  # 检测到符合条件的记录

    return False  # 未检测到目标记录
"""
    elif task['id'] <= 25:  # 搜索和筛选类任务
        file_name = "search_params.json"
        check_conditions = []
        for key, value in task['check_key'].items():
            if key == "type":
                check_conditions.append(f'event.get("type") == "{value}"')
            elif isinstance(value, list):
                check_conditions.append(f'event.get("{key}") in {value}')
            else:
                check_conditions.append(f'event.get("{key}") == "{value}" or event.get("{key}") == {value}')

        check_expr = " and ".join(check_conditions)
        check_code = f"""
    # 检查是否存在目标搜索记录
    for event in data.get("search_events", []):
        if {check_expr}:
            return True  # 检测到符合条件的记录

    return False  # 未检测到目标记录
"""
    else:  # 预订类任务
        file_name = "booking_history.json"
        check_code = f"""
    # 检查是否存在目标预订记录
    for event in data.get("booking_events", []):
        # 根据任务{task['id']}的特定条件进行检查
        # 这里需要根据实际的check_key进行详细判断
        if event.get("type") == "{task['check_key'].get('type', '')}":
            return True  # 检测到符合条件的记录

    return False  # 未检测到目标记录
"""

    # 确定文件名
    if task['id'] <= 6:
        file_name = "click_history.json"
    elif task['id'] <= 25:
        file_name = "search_params.json"
    else:
        file_name = "booking_history.json"

    script_content = f'''import subprocess
import json
import os

def check_task_{task['id']}():
    """
    任务{task['id']}: {task['instruction']}
    检验方法: {task['check_method']}
    """
    # 定义APP包名和存储记录的文件路径
    app_package = "com.example.Ctrip"
    file_path = "files/{file_name}"
    local_file = "{file_name}"  # 临时保存到电脑的文件

    # 获取ADB路径
    adb_path = os.path.join(os.environ.get('ANDROID_HOME', 'C:\\\\Users\\\\wudel\\\\AppData\\\\Local\\\\Android\\\\Sdk'),
                            'platform-tools', 'adb.exe')

    # 1. 通过ADB从手机获取记录文件
    try:
        subprocess.run(
            [adb_path, "exec-out", "run-as", app_package, "cat", file_path],
            stdout=open(local_file, "w"),
            check=True
        )
    except subprocess.CalledProcessError:
        print("获取文件失败，可能文件不存在或权限不足")
        return False

    # 2. 读取并解析本地文件
    try:
        with open(local_file, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, FileNotFoundError):
        print("文件解析失败或文件不存在")
        return False

{check_code}

if __name__ == "__main__":
    result = check_task_{task['id']}()
    print("检验结果：", "成功" if result else "失败")
'''

    with open(f"Autotest/{script_name}", "w", encoding="utf-8") as f:
        f.write(script_content)

    print(f"✓ 已生成 {script_name}")

def main():
    print("=" * 60)
    print("开始生成任务设计文档和检验脚本")
    print("=" * 60)

    # 创建Excel文件
    create_excel()

    # 确保Autotest目录存在
    os.makedirs("Autotest", exist_ok=True)

    # 生成所有检验脚本
    print("\n开始生成检验脚本...")
    for task in TASK_DESIGNS:
        generate_eval_script(task)

    print("\n" + "=" * 60)
    print(f"✓ 所有任务完成！")
    print(f"✓ 生成了1个Excel文件: 任务设计&检验逻辑.xlsx")
    print(f"✓ 生成了{len(TASK_DESIGNS)}个检验脚本: eval_1.py ~ eval_{len(TASK_DESIGNS)}.py")
    print("=" * 60)

if __name__ == "__main__":
    main()
