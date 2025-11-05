# coding: utf-8
import openpyxl

# 读取现有的Excel文件
wb = openpyxl.load_workbook('Autotest/任务设计&检验逻辑.xlsx')
ws = wb.active

print(f"工作表名称: {ws.title}")
print(f"最大行数: {ws.max_row}")
print(f"最大列数: {ws.max_column}")
print("\n前5行内容:")

for row in range(1, min(6, ws.max_row + 1)):
    row_data = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        row_data.append(str(cell.value)[:50] if cell.value else "")
    print(f"第{row}行: {row_data}")

print(f"\n列宽:")
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    if col in ws.column_dimensions:
        print(f"列{col}: {ws.column_dimensions[col].width}")
