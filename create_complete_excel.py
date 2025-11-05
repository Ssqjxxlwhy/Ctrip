# coding: utf-8
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "任务设计与检验逻辑"

# 设置列宽
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 70

# 创建样式
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 设置标题
headers = ['序号', '任务指令', '任务难度/页面', '检验方法', '检验脚本', '检验逻辑说明']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border_style

# 任务数据
tasks = [
    (1, '点击 "酒店" 图标，进入酒店预订页面。', '低 - 酒店预订页面', '点击记录检测',
     'eval_1.py', '在APP内部存储维护click_history.json文件。当用户点击"酒店"图标时，自动记录{time, icon:"酒店", page:"酒店预订页面"}。通过ADB直接读取该文件，检查click_events数组中是否存在匹配记录。'),

    (2, '点击 "机票" 图标，进入机票预订界面。', '低 - 机票预订界面', '点击记录检测',
     'eval_2.py', '在click_history.json中记录点击"机票"图标的操作，检查是否存在{icon:"机票", page:"机票预订界面"}的记录。'),

    (3, '点击 "火车票" 图标，进入火车票预订页面。', '低 - 火车票预订页面', '点击记录检测',
     'eval_3.py', '在click_history.json中记录点击"火车票"图标的操作，检查是否存在{icon:"火车票", page:"火车票预订页面"}的记录。'),

    (4, '点击"消息"按钮，进入消息页面', '低 - 消息页面', '点击记录检测',
     'eval_4.py', '在click_history.json中记录点击"消息"按钮的操作，检查是否存在{icon:"消息", page:"消息页面"}的记录。'),

    (5, '点击"行程"按钮，进入行程页面', '低 - 行程页面', '点击记录检测',
     'eval_5.py', '在click_history.json中记录点击"行程"按钮的操作，检查是否存在{icon:"行程", page:"行程页面"}的记录。'),

    (6, '点击"我的"按钮，进入我的页面', '低 - 我的页面', '点击记录检测',
     'eval_6.py', '在click_history.json中记录点击"我的"按钮的操作，检查是否存在{icon:"我的", page:"我的页面"}的记录。'),

    (7, '进入 "酒店" 页面，选择城市成都，得到酒店列表', '低 - 酒店列表页面', '搜索参数检测',
     'eval_7.py', '在search_params.json中记录搜索参数。通过ADB直接读取，检查search_events数组中是否存在{type:"hotel_search", city:"成都"}的记录。'),

    (8, '进入 "酒店" 页面，入住时间选择10月20日，退房时间选择10月21日，得到酒店列表', '低 - 酒店列表页面', '搜索参数检测',
     'eval_8.py', '在search_params.json中记录搜索参数，检查是否存在{type:"hotel_search", checkIn:"2025-10-20", checkOut:"2025-10-21"}的记录。'),

    (9, '进入 "酒店" 页面，选择间数1、成人数1、儿童数1，得到酒店列表', '低 - 酒店列表页面', '搜索参数检测',
     'eval_9.py', '在search_params.json中记录搜索参数，检查是否存在{type:"hotel_search", rooms:1, adults:1, children:1}的记录。'),

    (10, '进入 "机票" 页面，选出发地 "广州"、目的地 "深圳"，得到航班列表', '低 - 机票列表页面', '搜索参数检测',
     'eval_10.py', '在search_params.json中记录搜索参数，检查是否存在{type:"flight_search", from:"广州", to:"深圳"}的记录。'),

    (11, '进入 "机票" 页面，选择日期10月22日，得到航班列表', '低 - 机票列表页面', '搜索参数检测',
     'eval_11.py', '在search_params.json中记录搜索参数，检查是否存在{type:"flight_search", date:"2025-10-22"}的记录。'),

    (12, '进入"机票"页面，选择经济舱或者公务/头等舱，得到航班列表', '低 - 机票列表页面', '搜索参数检测',
     'eval_12.py', '在search_params.json中记录搜索参数，检查是否存在{type:"flight_search", cabin:["经济舱"或"公务舱"或"头等舱"]}的记录。'),

    (13, '进入 "火车票" 页面，选出发地 "北京"、目的地 "上海"，得到车次列表', '低 - 火车票列表页面', '搜索参数检测',
     'eval_13.py', '在search_params.json中记录搜索参数，检查是否存在{type:"train_search", from:"北京", to:"上海"}的记录。'),

    (14, '进入 "火车票" 页面，选择日期10月24日，得到车次列表', '低 - 火车票列表页面', '搜索参数检测',
     'eval_14.py', '在search_params.json中记录搜索参数，检查是否存在{type:"train_search", date:"2025-10-24"}的记录。'),

    (15, '进入 "火车票" 页面，选择学生票，得到车次列表', '低 - 火车票列表页面', '搜索参数检测',
     'eval_15.py', '在search_params.json中记录搜索参数，检查是否存在{type:"train_search", ticketType:"学生票"}的记录。'),

    (16, '进入 "酒店" 页面，城市选北京，入住时间选择10月22日，退房时间选择10月23日，得到酒店列表', '中 - 酒店列表页面', '搜索参数检测',
     'eval_16.py', '在search_params.json中记录搜索参数，检查是否存在{type:"hotel_search", city:"北京", checkIn:"2025-10-22", checkOut:"2025-10-23"}的记录。'),

    (17, '进入 "酒店" 页面，城市选上海，入住时间今天，退房时间明天，间数选择2，成人数选择2，儿童数选择0，得到酒店列表', '中 - 酒店列表页面', '搜索参数检测',
     'eval_17.py', '在search_params.json中记录搜索参数，检查是否存在{type:"hotel_search", city:"上海", rooms:2, adults:2, children:0}的记录（日期动态判断）。'),

    (18, '进入 "机票" 页面，选出发地 "北京"、目的地 "深圳"，选择日期10月25号，得到航班列表', '中 - 机票列表页面', '搜索参数检测',
     'eval_18.py', '在search_params.json中记录搜索参数，检查是否存在{type:"flight_search", from:"北京", to:"深圳", date:"2025-10-25"}的记录。'),

    (19, '进入 "机票" 页面，选出发地 "成都"、目的地 "上海"，选择日期10月20号，选头等舱，得到航班列表', '中 - 机票列表页面', '搜索参数检测',
     'eval_19.py', '在search_params.json中记录搜索参数，检查是否存在{type:"flight_search", from:"成都", to:"上海", date:"2025-10-20", cabin:"头等舱"}的记录。'),

    (20, '进入 "火车票" 页面，选出发地 "北京"、目的地 "上海"，选择日期10月22日，得到车次列表', '中 - 火车票列表页面', '搜索参数检测',
     'eval_20.py', '在search_params.json中记录搜索参数，检查是否存在{type:"train_search", from:"北京", to:"上海", date:"2025-10-22"}的记录。'),

    (21, '进入 "火车票" 页面，选出发地 "长沙"、目的地 "天津"，选择日期1月18日，选择学生票，得到车次列表', '中 - 火车票列表页面', '搜索参数检测',
     'eval_21.py', '在search_params.json中记录搜索参数，检查是否存在{type:"train_search", from:"长沙", to:"天津", date:"2026-01-18", ticketType:"学生票"}的记录。'),

    (22, '进入 "酒店" 页面，为我筛选北京评分最高的前 5 家酒店', '中 - 酒店列表页面', '搜索+筛选检测',
     'eval_22.py', '在search_params.json中记录搜索和筛选参数，检查是否存在{type:"hotel_search", city:"北京", sortBy:"rating", limit:5}的记录。'),

    (23, '进入 "酒店" 页面，为我检索上海价格最低的酒店', '中 - 酒店列表页面', '搜索+筛选检测',
     'eval_23.py', '在search_params.json中记录搜索和筛选参数，检查是否存在{type:"hotel_search", city:"上海", sortBy:"price_asc"}的记录。'),

    (24, '进入 "机票" 页面，查10月21日从北京飞广州的机票，统计下最便宜的 3 趟航班的平均价格。', '中 - 机票列表页面', '搜索+计算检测',
     'eval_24.py', '在search_params.json中记录搜索和计算结果，检查是否存在{type:"flight_search", from:"北京", to:"广州", date:"2025-10-21", calculation:"average_price_top3"}的记录。'),

    (25, '进入 "火车票" 页面，查10月22日广州到杭州下午2点到5点的车次，得到车次列表。', '中 - 火车票列表页面', '搜索+时间筛选检测',
     'eval_25.py', '在search_params.json中记录搜索和时间筛选，检查是否存在{type:"train_search", from:"广州", to:"杭州", date:"2025-10-22", timeRange:"14:00-17:00"}的记录。'),

    (26, '选上海的酒店，入住日期选10月22日，退房10月25日，客房数和入住人数默认，得到酒店列表后，点击第一个酒店，然后选择第一个房型预订', '中 - 酒店预订', '预订记录检测',
     'eval_26.py', '在booking_history.json中记录预订操作。通过ADB直接读取，检查booking_events数组中是否存在{type:"hotel_booking", city:"上海", checkIn:"2025-10-22", checkOut:"2025-10-25", hotelIndex:0, roomIndex:0}的记录。'),

    (27, '进入 "机票" 页面，选出发地 "武汉"、目的地 "深圳"，选择日期11月10号，舱型默认，得到航班列表后，预订第一架航班', '中 - 机票预订', '预订记录检测',
     'eval_27.py', '在booking_history.json中记录预订操作，检查是否存在{type:"flight_booking", from:"武汉", to:"深圳", date:"2025-11-10", flightIndex:0}的记录。'),

    (28, '进入 "火车票" 页面，选出发地 "北京"、目的地 "上海"，选择日期10月23日，得到车次列表后，预订第一班车次', '中 - 火车票预订', '预订记录检测',
     'eval_28.py', '在booking_history.json中记录预订操作，检查是否存在{type:"train_booking", from:"北京", to:"上海", date:"2025-10-23", trainIndex:0}的记录。'),

    (29, '选上海的酒店，入住日期选10月21日，退房10月25日，客房数和入住人数默认，得到酒店列表后，选择价格最低的酒店，然后选择最便宜的房型预订', '中 - 酒店预订', '预订记录检测',
     'eval_29.py', '在booking_history.json中记录预订操作，检查是否存在{type:"hotel_booking", city:"上海", checkIn:"2025-10-21", checkOut:"2025-10-25", selection:"cheapest"}的记录。'),

    (30, '进入 "火车票" 页面，选出发地 "北京"、目的地 "上海"，选择日期10月20日，得到车次列表后，预订下午3点到下午4点的任意一班车次', '中 - 火车票预订', '预订记录检测',
     'eval_30.py', '在booking_history.json中记录预订操作，检查是否存在{type:"train_booking", from:"北京", to:"上海", date:"2025-10-20", timeRange:"15:00-16:00"}的记录。'),

    (31, '先订北京飞上海的机票（10月20日，经济舱），再订上海的酒店（入住10月20日，退房10月21日），最后订上海回北京的火车票（10月21日）', '高 - 多步骤任务', '多步骤预订检测',
     'eval_31.py', '在booking_history.json中记录多步骤预订。检查是否按顺序完成：1)机票(北京->上海,经济舱) 2)酒店(上海,10.20-21) 3)火车票(上海->北京)。验证type:"multi_booking"且包含3个步骤。'),

    (32, '订 10 月 20 日从杭州到北京的最快火车票（5 小时内达），住北京中关村亚朵酒店两晚（10.20 - 10.22），再订 10.22 北京回杭州的最晚高铁，计算所有费用后判断 2000 元够不够。', '高 - 复杂任务', '复杂任务+计算检测',
     'eval_32.py', '在booking_history.json中记录复杂任务。检查是否完成：1)火车票(杭州->北京,<=5h) 2)酒店(中关村亚朵) 3)火车票(北京->杭州,最晚) 4)费用计算及预算判断。验证包含totalCost和budgetCheck字段。'),

    (33, '订 5 张 10 月 20 日深圳到北京的火车票（要求 5 小时内到达）', '高 - 批量任务', '批量预订检测',
     'eval_33.py', '在booking_history.json中记录批量预订。检查是否存在{type:"batch_booking", bookingType:"train", from:"深圳", to:"北京", date:"2025-10-20", quantity:5, constraint:"duration<=5h"}的记录。'),

    (34, '订 5 张 10 月 20 日广州到北京的最早火车票，再订北京 3 家不同的酒店', '高 - 复合批量任务', '复合批量预订检测',
     'eval_34.py', '在booking_history.json中记录复合批量预订。检查是否完成：1)5张火车票(广州->北京,最早) 2)3家不同酒店。验证type:"complex_batch_booking"。'),

    (35, '订 10 月 15 日北京飞上海的 6 张经济舱机票', '高 - 批量任务', '批量预订检测',
     'eval_35.py', '在booking_history.json中记录批量预订。检查是否存在{type:"batch_booking", bookingType:"flight", from:"北京", to:"上海", date:"2025-10-15", cabin:"经济舱", quantity:6}的记录。')
]

# 填充数据并应用样式
cell_alignment = Alignment(wrap_text=True, vertical='top')
center_alignment = Alignment(horizontal='center', vertical='center')

for task_id, instruction, difficulty, method, script, logic in tasks:
    row = task_id + 1

    # 填充数据
    ws.cell(row=row, column=1, value=task_id)
    ws.cell(row=row, column=2, value=instruction)
    ws.cell(row=row, column=3, value=difficulty)
    ws.cell(row=row, column=4, value=method)
    ws.cell(row=row, column=5, value=script)
    ws.cell(row=row, column=6, value=logic)

    # 应用样式
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.border = border_style
        if col == 1 or col == 5:  # 序号和脚本列居中
            cell.alignment = center_alignment
        else:
            cell.alignment = cell_alignment

# 添加说明行
note_row = len(tasks) + 3
ws.merge_cells(f'A{note_row}:F{note_row}')
note_cell = ws.cell(row=note_row, column=1)
note_cell.value = '说明：所有检验脚本位于Autotest目录，通过ADB直接读取手机内部存储的JSON文件进行检验，无需保存临时文件。'
note_cell.font = Font(italic=True, size=10, color='666666')
note_cell.alignment = Alignment(wrap_text=True)

# 添加JSON文件说明
json_note_row = note_row + 1
ws.merge_cells(f'A{json_note_row}:F{json_note_row}')
json_note_cell = ws.cell(row=json_note_row, column=1)
json_note_cell.value = 'JSON文件：click_history.json(任务1-6) | search_params.json(任务7-25) | booking_history.json(任务26-35)'
json_note_cell.font = Font(italic=True, size=10, color='666666')
json_note_cell.alignment = Alignment(wrap_text=True)

# 保存文件
wb.save('任务设计&检验逻辑.xlsx')
print('Excel file created successfully!')
print('Contains 35 tasks with complete design and verification logic')
print('Filename: 任务设计&检验逻辑.xlsx')
