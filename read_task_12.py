import openpyxl

wb = openpyxl.load_workbook('Autotest/任务设计&检验逻辑.xlsx')
ws = wb.active

# 第12条指令在第13行
print(f"第12条指令（第13行）:")
print(f"指令内容: {ws.cell(13, 2).value}")  # 第2列是指令内容
print(f"检验方法: {ws.cell(13, 3).value}")  # 第3列是检验方法
