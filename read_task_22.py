import openpyxl

wb = openpyxl.load_workbook('Autotest/任务设计&检验逻辑.xlsx')
ws = wb.active

# 第22条指令在第23行
row = 23
print(f"第22条指令：")
print(f"序号: {ws.cell(row, 1).value}")
print(f"指令内容: {ws.cell(row, 2).value}")
print(f"检验方法: {ws.cell(row, 3).value}")
