from openpyxl import load_workbook

# 读取 Excel 文件
wb = load_workbook('Autotest/任务设计&检验逻辑.xlsx')
ws = wb.active

# 打印所有内容
for row in ws.iter_rows(values_only=True):
    print(row)
