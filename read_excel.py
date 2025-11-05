import openpyxl

try:
    # 读取Excel文件
    wb = openpyxl.load_workbook('任务设计-脑暴.xlsx')

    # 将输出写入文件
    with open('excel_content.txt', 'w', encoding='utf-8') as f:
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            f.write(f"\n{'='*60}\n")
            f.write(f"工作表: {sheet_name}\n")
            f.write('='*60 + '\n')

            ws = wb[sheet_name]

            # 读取所有行
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                # 过滤掉全空行
                if any(cell is not None for cell in row):
                    f.write(f"行{row_idx}: {row}\n")

    wb.close()
    print("内容已保存到 excel_content.txt")

except FileNotFoundError:
    print("错误: 找不到文件 '任务设计-脑暴.xlsx'")
except Exception as e:
    print(f"错误: {e}")
    import traceback
    traceback.print_exc()
