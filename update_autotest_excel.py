import openpyxl
import sys

# Excel文件路径
excel_path = r"Autotest\任务设计&检验逻辑.xlsx"

try:
    # 打开Excel文件
    print(f"正在打开Excel文件: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    print("Excel文件打开成功！")

    # 更新任务22的检验方法和检验脚本
    # 任务22在第23行（行号从1开始）
    print("\n更新任务22...")
    row_22 = 23
    ws.cell(row=row_22, column=3).value = """1. 计算真实答案：从北京酒店数据中按评分降序排序，取前3家，计算这3家的平均价格
   - 按评分排序：旭阳趣舍(4.8, ¥65) → 一家人(4.7, ¥104) → 希尔顿(4.6, ¥380)
   - 平均价格：(65 + 104 + 380) / 3 = 183元
2. 智能体答案：接收智能体返回的平均价格（可以是数字、字符串或包含价格的字典）
3. 对比判断：允许±1元的误差（考虑四舍五入）
4. 辅助检查：也可从search_params.json检查是否记录了{type:"hotel_search", city:"北京", sortBy:"rating", limit:3, calculation:"average_price"}"""
    ws.cell(row=row_22, column=4).value = "check_hotel_search_beijing_top3_avg_price(agent_answer)"

    # 更新任务23的检验方法和检验脚本
    # 任务23在第24行
    print("更新任务23...")
    row_23 = 24
    ws.cell(row=row_23, column=3).value = """1. 计算真实答案：从上海酒店数据中找出所有酒店的最低价格
   - 所有酒店价格：180, 220, 298, 350, 888
   - 最低价格：180元（对应上海田子坊创意民宿）
2. 智能体答案：接收智能体返回的价格（可以是数字、字符串或包含价格的字典）
3. 对比判断：精确匹配，必须等于180元
4. 辅助检查：也可从search_params.json检查是否记录了{type:"hotel_search", city:"上海", calculation:"min_price"}"""
    ws.cell(row=row_23, column=4).value = "check_hotel_search_shanghai_min_price(agent_answer)"

    # 任务24保持不变
    print("任务24保持不变...")

    # 更新任务25的检验方法和检验脚本
    # 任务25在第26行
    print("更新任务25...")
    row_25 = 26
    ws.cell(row=row_25, column=3).value = """1. 计算真实答案：
   - 从trains_list.json加载火车票数据
   - 筛选广州到杭州10月22日下午2点到5点（14:00-17:00）的车次
   - 找出这些车次中所有座位类型的价格，取最高价格
   - 真实答案：K535车次的商务座 1410.0元（如果数据不同，以实际为准）
2. 智能体答案：接收智能体返回的价格（可以是数字、字符串或包含价格的字典）
3. 对比判断：允许±1元的误差
4. 辅助检查：也可从search_params.json检查是否记录了{type:"train_search", from:"广州", to:"杭州", date:"2025-10-22", timeRange:"14:00-17:00", calculation:"max_price"}"""
    ws.cell(row=row_25, column=4).value = "check_train_search_max_price(agent_answer)"

    # 保存文件
    print("\n正在保存Excel文件...")
    wb.save(excel_path)
    print("Excel文件更新成功！")

    # 打印更新内容
    print("\n" + "=" * 80)
    print("更新内容摘要：")
    print("=" * 80)
    print(f"任务22 - 检验函数：check_hotel_search_beijing_top3_avg_price(agent_answer)")
    print(f"         检验方法：计算北京评分最高前3家酒店的平均价格(183元)，对比智能体答案")
    print()
    print(f"任务23 - 检验函数：check_hotel_search_shanghai_min_price(agent_answer)")
    print(f"         检验方法：计算上海所有酒店的最低价格(180元)，对比智能体答案")
    print()
    print(f"任务24 - 保持不变：check_flight_search_price_avg(agent_answer)")
    print(f"         检验方法：计算北京到广州最便宜3趟航班的平均价格(625元)，对比智能体答案")
    print()
    print(f"任务25 - 检验函数：check_train_search_max_price(agent_answer)")
    print(f"         检验方法：计算下午2-5点车次中最高价格(1410元)，对比智能体答案")
    print("=" * 80)

except PermissionError:
    print("错误：Excel文件可能正在被其他程序打开。")
    print("请关闭Excel文件后重新运行此脚本。")
    sys.exit(1)
except Exception as e:
    print(f"错误：{e}")
    sys.exit(1)
