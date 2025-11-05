# coding: utf-8
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "任务设计与检验逻辑"

# 设置列宽
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 100

# 创建标题行样式
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_alignment = Alignment(horizontal='center', vertical='center')

# 设置标题
headers = ['序号', '任务指令', '检验方法', '检验逻辑说明']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# 任务数据（避免使用中文引号）
tasks = [
    (1, '点击 "酒店" 图标，进入酒店预订页面', '维护点击记录存储',
     '在手机APP的内部存储中，维护一个记录用户操作历史的JSON文件（click_history.json），每当用户点击"酒店"图标时，自动将点击事件的相关信息（如点击时间、图标名称、跳转页面）写入该文件。通过ADB指令从手机中获取该JSON文件，解析文件内容，检查是否存在icon="酒店"且page="酒店预订页面"的记录。'),

    (2, '点击 "机票" 图标，进入机票预订界面', '维护点击记录存储',
     '在手机APP的内部存储中维护click_history.json文件，每当用户点击"机票"图标时，自动将点击事件写入该文件。通过ADB指令获取并解析JSON文件，检查是否存在icon="机票"且page="机票预订界面"的记录。'),

    (3, '点击 "火车票" 图标，进入火车票预订页面', '维护点击记录存储',
     '在手机APP的内部存储中维护click_history.json文件，每当用户点击"火车票"图标时，自动将点击事件写入该文件。通过ADB指令获取并解析JSON文件，检查是否存在icon="火车票"且page="火车票预订页面"的记录。'),

    (4, '点击"消息"按钮，进入消息页面', '维护点击记录存储',
     '在click_history.json中记录点击"消息"按钮的操作，通过ADB获取文件，检查是否存在icon="消息"且page="消息页面"的记录。'),

    (5, '点击"行程"按钮，进入行程页面', '维护点击记录存储',
     '在click_history.json中记录点击"行程"按钮的操作，通过ADB获取文件，检查是否存在icon="行程"且page="行程页面"的记录。'),

    (6, '点击"我的"按钮，进入我的页面', '维护点击记录存储',
     '在click_history.json中记录点击"我的"按钮的操作，通过ADB获取文件，检查是否存在icon="我的"且page="我的页面"的记录。'),

    (7, '进入 "酒店" 页面，选择城市成都，得到酒店列表', '维护搜索参数记录存储',
     '在APP内部存储中维护搜索参数记录JSON文件（search_params.json），当用户在酒店页面选择城市后进行搜索时，记录搜索参数（type, city, date等）。通过ADB获取该文件，检查search_events数组中是否存在type="hotel_search"且city="成都"的记录。'),

    (8, '进入 "酒店" 页面，入住时间选择10月20日，退房时间选择10月21日，得到酒店列表', '维护搜索参数记录存储',
     '在search_params.json中记录酒店搜索参数，通过ADB获取文件，检查是否存在type="hotel_search"、checkIn="2025-10-20"、checkOut="2025-10-21"的搜索记录。'),

    (9, '进入 "酒店" 页面，选择间数1、成人数1、儿童数1，得到酒店列表', '维护搜索参数记录存储',
     '在search_params.json中记录酒店搜索参数，检查是否存在type="hotel_search"、rooms=1、adults=1、children=1的搜索记录。'),

    (10, '进入 "机票" 页面，选出发地 "广州"、目的地 "深圳"，得到航班列表', '维护搜索参数记录存储',
     '在search_params.json中记录机票搜索参数，检查是否存在type="flight_search"、from="广州"、to="深圳"的搜索记录。'),

    (11, '进入 "机票" 页面，选择日期10月22日，得到航班列表', '维护搜索参数记录存储',
     '在search_params.json中记录机票搜索参数，检查是否存在type="flight_search"、date="2025-10-22"的搜索记录。'),

    (12, '进入"机票"页面，选择经济舱或者公务/头等舱，得到航班列表', '维护搜索参数记录存储',
     '在search_params.json中记录机票搜索参数，检查是否存在type="flight_search"且cabin字段为"经济舱"或"公务舱"或"头等舱"的搜索记录。'),

    (13, '进入 "火车票" 页面，选出发地 "北京"、目的地 "上海"，得到车次列表', '维护搜索参数记录存储',
     '在search_params.json中记录火车票搜索参数，检查是否存在type="train_search"、from="北京"、to="上海"的搜索记录。'),

    (14, '进入 "火车票" 页面，选择日期10月24日，得到车次列表', '维护搜索参数记录存储',
     '在search_params.json中记录火车票搜索参数，检查是否存在type="train_search"、date="2025-10-24"的搜索记录。'),

    (15, '进入 "火车票" 页面，选择学生票，得到车次列表', '维护搜索参数记录存储',
     '在search_params.json中记录火车票搜索参数，检查是否存在type="train_search"、ticketType="学生票"的搜索记录。'),

    (16, '进入 "酒店" 页面，城市选北京，入住时间选择10月22日，退房时间选择10月23日，得到酒店列表', '维护搜索参数记录存储',
     '在search_params.json中记录酒店搜索参数，检查是否存在type="hotel_search"、city="北京"、checkIn="2025-10-22"、checkOut="2025-10-23"的搜索记录。'),

    (17, '进入 "酒店" 页面，城市选上海，入住时间今天，退房时间明天，间数选择2，成人数选择2，儿童数选择0，得到酒店列表', '维护搜索参数记录存储',
     '在search_params.json中记录酒店搜索参数，检查是否存在type="hotel_search"、city="上海"、rooms=2、adults=2、children=0的搜索记录（日期动态判断为今明两天）。'),

    (18, '进入 "机票" 页面，选出发地 "北京"、目的地 "深圳"，选择日期10月25号，得到航班列表', '维护搜索参数记录存储',
     '在search_params.json中记录机票搜索参数，检查是否存在type="flight_search"、from="北京"、to="深圳"、date="2025-10-25"的搜索记录。'),

    (19, '进入 "机票" 页面，选出发地 "成都"、目的地 "上海"，选择日期10月20号，选头等舱，得到航班列表', '维护搜索参数记录存储',
     '在search_params.json中记录机票搜索参数，检查是否存在type="flight_search"、from="成都"、to="上海"、date="2025-10-20"、cabin="头等舱"的搜索记录。'),

    (20, '进入 "火车票" 页面，选出发地 "北京"、目的地 "上海"，选择日期10月22日，得到车次列表', '维护搜索参数记录存储',
     '在search_params.json中记录火车票搜索参数，检查是否存在type="train_search"、from="北京"、to="上海"、date="2025-10-22"的搜索记录。'),

    (21, '进入 "火车票" 页面，选出发地 "长沙"、目的地 "天津"，选择日期1月18日，选择学生票，得到车次列表', '维护搜索参数记录存储',
     '在search_params.json中记录火车票搜索参数，检查是否存在type="train_search"、from="长沙"、to="天津"、date="2026-01-18"、ticketType="学生票"的搜索记录。'),

    (22, '进入 "酒店" 页面，为我筛选北京评分最高的前 5 家酒店', '维护搜索参数和筛选条件记录',
     '在search_params.json中记录酒店搜索和筛选条件，检查是否存在type="hotel_search"、city="北京"、sortBy="rating"、limit=5的记录。'),

    (23, '进入 "酒店" 页面，为我检索上海价格最低的酒店', '维护搜索参数和筛选条件记录',
     '在search_params.json中记录酒店搜索和筛选条件，检查是否存在type="hotel_search"、city="上海"、sortBy="price_asc"的记录。'),

    (24, '进入 "机票" 页面，查10月21日从北京飞广州的机票，统计下最便宜的 3 趟航班的平均价格', '维护搜索参数和计算结果记录',
     '在search_params.json中记录机票搜索和计算结果，检查是否存在type="flight_search"、from="北京"、to="广州"、date="2025-10-21"且包含最便宜3趟航班平均价格计算结果的记录。'),

    (25, '进入 "火车票" 页面，查10月22日广州到杭州下午2点到5点的车次，得到车次列表', '维护搜索参数和时间筛选记录',
     '在search_params.json中记录火车票搜索和时间筛选，检查是否存在type="train_search"、from="广州"、to="杭州"、date="2025-10-22"、timeRange="14:00-17:00"的记录。'),

    (26, '选上海的酒店，入住日期选10月22日，退房10月25日，客房数和入住人数默认，得到酒店列表后，点击第一个酒店，然后选择第一个房型预订', '维护预订记录存储',
     '在booking_history.json中记录预订操作，检查booking_events数组中是否存在type="hotel_booking"、city="上海"、checkIn="2025-10-22"、checkOut="2025-10-25"、hotelIndex=0、roomIndex=0的预订记录。'),

    (27, '进入 "机票" 页面，选出发地 "武汉"、目的地 "深圳"，选择日期11月10号，舱型默认，得到航班列表后，预订第一架航班', '维护预订记录存储',
     '在booking_history.json中记录机票预订，检查是否存在type="flight_booking"、from="武汉"、to="深圳"、date="2025-11-10"、flightIndex=0的预订记录。'),

    (28, '进入 "火车票" 页面，选出发地 "北京"、目的地 "上海"，选择日期10月23日，得到车次列表后，预订第一班车次', '维护预订记录存储',
     '在booking_history.json中记录火车票预订，检查是否存在type="train_booking"、from="北京"、to="上海"、date="2025-10-23"、trainIndex=0的预订记录。'),

    (29, '选上海的酒店，入住日期选10月21日，退房10月25日，客房数和入住人数默认，得到酒店列表后，选择价格最低的酒店，然后选择最便宜的房型预订', '维护预订记录存储',
     '在booking_history.json中记录酒店预订，检查是否存在type="hotel_booking"、city="上海"、checkIn="2025-10-21"、checkOut="2025-10-25"、selection="cheapest"的预订记录。'),

    (30, '进入 "火车票" 页面，选出发地 "北京"、目的地 "上海"，选择日期10月20日，得到车次列表后，预订下午3点到下午4点的任意一班车次', '维护预订记录存储',
     '在booking_history.json中记录火车票预订，检查是否存在type="train_booking"、from="北京"、to="上海"、date="2025-10-20"、timeRange="15:00-16:00"的预订记录。'),

    (31, '先订北京飞上海的机票（10月20日，经济舱），再订上海的酒店（入住10月20日，退房10月21日），最后订上海回北京的火车票（10月21日）', '维护多步骤任务记录',
     '在booking_history.json中记录多步骤预订，检查是否按顺序完成三个预订：1)机票(北京->上海,2025-10-20,经济舱) 2)酒店(上海,2025-10-20至21) 3)火车票(上海->北京,2025-10-21)，并验证顺序正确。'),

    (32, '订 10 月 20 日从杭州到北京的最快火车票（5 小时内达），住北京中关村亚朵酒店两晚（10.20 - 10.22），再订 10.22 北京回杭州的最晚高铁，计算所有费用后判断 2000 元够不够', '维护多步骤任务和计算结果记录',
     '在booking_history.json中记录多步骤预订和费用计算，检查是否完成：1)火车票(杭州->北京,<=5h) 2)酒店(中关村亚朵,10.20-22) 3)火车票(北京->杭州,最晚) 4)总费用计算及2000元预算判断。'),

    (33, '订 5 张 10 月 20 日深圳到北京的火车票（要求 5 小时内到达）', '维护批量预订记录',
     '在booking_history.json中记录批量预订，检查是否存在5张火车票记录：from="深圳"、to="北京"、date="2025-10-20"、duration<=5h。'),

    (34, '订 5 张 10 月 20 日广州到北京的最早火车票，再订北京 3 家不同的酒店', '维护批量和多类型预订记录',
     '在booking_history.json中记录复合批量预订，检查是否完成：1)5张火车票(广州->北京,2025-10-20,最早) 2)3家不同的北京酒店预订。'),

    (35, '订 10 月 15 日北京飞上海的 6 张经济舱机票', '维护批量预订记录',
     '在booking_history.json中记录批量预订，检查是否存在6张机票记录：from="北京"、to="上海"、date="2025-10-15"、cabin="经济舱"。')
]

# 填充数据
for task_id, instruction, method, logic in tasks:
    row = task_id + 1
    ws.cell(row=row, column=1, value=task_id)
    ws.cell(row=row, column=2, value=instruction)
    ws.cell(row=row, column=3, value=method)
    ws.cell(row=row, column=4, value=logic)

    # 设置单元格样式
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        if col == 1:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# 保存文件
wb.save('任务设计&检验逻辑.xlsx')
print('Excel文件已创建成功!')
